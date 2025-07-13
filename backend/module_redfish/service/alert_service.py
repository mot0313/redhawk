"""
告警管理Service层
"""
import math
from datetime import datetime
from typing import List, Optional, Dict, Any
from sqlalchemy.ext.asyncio import AsyncSession
from module_redfish.dao.alert_dao import AlertDao
from module_redfish.entity.vo.alert_vo import (
    AlertPageQueryModel,
    AlertStatisticsModel, AlertTrendModel, RealtimeAlertModel, ScheduledAlertModel,
    AlertDistributionModel, MaintenanceScheduleModel, MaintenanceUpdateModel, 
    BatchMaintenanceUpdateModel, MaintenancePageQueryModel
)
from utils.response_util import ResponseUtil
from utils.page_util import PageResponseModel, PageUtil
from utils.log_util import logger


class AlertService:
    """告警管理服务"""
    
    @classmethod
    async def get_alert_list_services(
        cls,
        db: AsyncSession,
        query_object: AlertPageQueryModel,
        is_page: bool = False
    ) -> PageResponseModel:
        """
        获取告警列表
        
        Args:
            db: 数据库会话
            query_object: 查询对象
            is_page: 是否分页
            
        Returns:
            PageResponseModel: 分页响应
        """
        alert_list, total = await AlertDao.get_alert_list(db, query_object, is_page)
        
        if is_page:
            # 使用AlertPageResponseModel创建分页响应
            from module_redfish.entity.vo.alert_vo import AlertPageResponseModel
            return AlertPageResponseModel.create(
                alerts=alert_list,
                page_num=query_object.page_num,
                page_size=query_object.page_size,
                total=total
            )
        else:
            # 非分页模式，返回告警列表
            from module_redfish.entity.vo.alert_vo import AlertResponseModel
            alert_models = []
            for alert in alert_list:
                if hasattr(alert, '__dict__'):
                    alert_dict = alert.__dict__.copy()
                    alert_dict.pop('_sa_instance_state', None)
                else:
                    alert_dict = alert
                alert_models.append(AlertResponseModel(**alert_dict))
            return alert_models
    
    @classmethod
    async def get_alert_detail_services(cls, db: AsyncSession, alert_id: int):
        """
        获取告警详情
        
        Args:
            db: 数据库会话
            alert_id: 告警ID
            
        Returns:
            AlertDetailResponseModel: 响应结果
        """
        from module_redfish.entity.vo.alert_vo import AlertDetailResponseModel
        alert = await AlertDao.get_alert_by_id(db, alert_id)
        if not alert:
            return None
        
        return AlertDetailResponseModel.create(alert=alert)
    
    # 精简版移除手动解决和忽略告警功能，告警状态由监控系统自动管理
    
    @classmethod
    async def get_alert_statistics_services(cls, db: AsyncSession, days: int = 7):
        """
        获取告警统计信息
        
        Args:
            db: 数据库会话
            days: 统计天数
            
        Returns:
            AlertStatsResponseModel: 统计信息
        """
        from module_redfish.entity.vo.alert_vo import AlertStatsResponseModel
        stats = await AlertDao.get_alert_statistics(db, days)
        
        return AlertStatsResponseModel.create(stats)
    
    @classmethod
    async def get_alert_trend_services(cls, db: AsyncSession, days: int = 7) -> List[AlertTrendModel]:
        """
        获取告警趋势数据
        
        Args:
            db: 数据库会话
            days: 统计天数
            
        Returns:
            List[AlertTrendModel]: 趋势数据
        """
        trend_data = await AlertDao.get_alert_trend(db, days)
        
        return [
            AlertTrendModel(
                date=item['date'],
                urgent_count=item['urgent_count'],
                scheduled_count=item['scheduled_count'],
                total_count=item['total_count']
            )
            for item in trend_data
        ]
    
    @classmethod
    async def get_realtime_alerts_services(cls, db: AsyncSession, limit: int = 10) -> List[RealtimeAlertModel]:
        """
        获取实时告警列表
        
        Args:
            db: 数据库会话
            limit: 返回数量限制
            
        Returns:
            List[RealtimeAlertModel]: 实时告警列表
        """
        alerts = await AlertDao.get_realtime_alerts(db, limit)
        
        return [
            RealtimeAlertModel(
                device_id=alert['device_id'],
                hostname=alert['hostname'],
                business_ip=alert['business_ip'],
                component_type=alert['component_type'],
                component_name=alert['component_name'],
                health_status=alert['health_status'],
                urgency_level=alert['urgency_level'],
                last_occurrence=alert['last_occurrence']
            )
            for alert in alerts
        ]
    
    @classmethod
    async def get_scheduled_alerts_services(cls, db: AsyncSession, limit: int = 10) -> List[ScheduledAlertModel]:
        """
        获取择期告警列表
        
        Args:
            db: 数据库会话
            limit: 返回数量限制
            
        Returns:
            List[ScheduledAlertModel]: 择期告警列表
        """
        alerts = await AlertDao.get_scheduled_alerts(db, limit)
        
        return [
            ScheduledAlertModel(
                device_id=alert['device_id'],
                hostname=alert['hostname'],
                business_ip=alert['business_ip'],
                component_type=alert['component_type'],
                component_name=alert['component_name'],
                health_status=alert['health_status'],
                alert_message="",  # 优化版DAO中没有alert_message字段
                first_occurrence=alert['first_occurrence']
            )
            for alert in alerts
        ]
    
    @classmethod
    async def get_alert_distribution_services(cls, db: AsyncSession) -> AlertDistributionModel:
        """
        获取告警分布统计
        
        Args:
            db: 数据库会话
            
        Returns:
            AlertDistributionModel: 分布统计
        """
        distribution = await AlertDao.get_alert_distribution(db)
        
        return AlertDistributionModel(
            byLevel=distribution['by_level'],
            byComponent=distribution['by_component'],
            byLocation=distribution['by_location'],
            byManufacturer=distribution['by_manufacturer']
        )
    
    @classmethod
    async def create_or_update_alert_services(
        cls,
        db: AsyncSession,
        device_id: int,
        component_type: str,
        component_name: str,
        health_status: str,
        alert_message: str,
        urgency_level: str
    ) -> Optional[int]:
        """
        创建或更新告警（由监控系统调用）
        
        Args:
            db: 数据库会话
            device_id: 设备ID
            component_type: 组件类型
            component_name: 组件名称
            health_status: 健康状态
            alert_message: 告警消息
            urgency_level: 紧急程度
            
        Returns:
            Optional[int]: 告警ID
        """
        try:
            alert = await AlertDao.get_or_create_alert(
                db,
                device_id,
                component_type,
                component_name,
                health_status,
                alert_message,
                alert_level
            )
            return alert.alert_id
        except Exception as e:
            logger.error(f"创建或更新告警失败: {str(e)}")
            return None
    
    @classmethod
    async def get_device_alerts_services(cls, db: AsyncSession, device_id: int) -> List[Dict[str, Any]]:
        """
        获取设备的所有活跃告警
        
        Args:
            db: 数据库会话
            device_id: 设备ID
            
        Returns:
            List[Dict[str, Any]]: 告警列表
        """
        query_object = AlertPageQueryModel(
            device_id=device_id,
            status='active',
            page_num=0,
            page_size=1000
        )
        
        alerts, _ = await AlertDao.get_alert_list(db, query_object, is_page=False)
        
        return [
            {
                'alert_id': alert.alert_id,
                'component_type': alert.component_type,
                'component_name': alert.component_name,
                'health_status': alert.health_status,
                'urgency_level': alert.urgency_level,
                'alert_message': alert.alert_message,
                'first_occurrence': alert.first_occurrence,
                'last_occurrence': alert.last_occurrence
            }
            for alert in alerts
        ]

    @classmethod
    async def schedule_maintenance_services(
        cls,
        db: AsyncSession,
        maintenance_schedule: MaintenanceScheduleModel
    ) -> ResponseUtil:
        """
        为告警安排维修时间
        
        Args:
            db: 数据库会话
            maintenance_schedule: 维修计划模型
            
        Returns:
            ResponseUtil: 响应结果
        """
        try:
            maintenance_data = {
                'scheduled_maintenance_time': maintenance_schedule.scheduled_maintenance_time,
                'maintenance_description': maintenance_schedule.maintenance_description,
                'maintenance_notes': maintenance_schedule.maintenance_notes
            }
            
            success = await AlertDao.schedule_maintenance(
                db, 
                maintenance_schedule.alert_id, 
                maintenance_data
            )
            
            if success:
                logger.info(f"成功为告警 {maintenance_schedule.alert_id} 安排维修时间")
                return ResponseUtil.success(msg="安排维修时间成功")
            else:
                return ResponseUtil.failure(msg="告警不存在或安排维修时间失败")
                
        except Exception as e:
            logger.error(f"安排维修时间失败: {str(e)}")
            return ResponseUtil.failure(msg="安排维修时间失败")

    @classmethod
    async def update_maintenance_services(
        cls,
        db: AsyncSession,
        maintenance_update: MaintenanceUpdateModel
    ) -> ResponseUtil:
        """
        更新告警维修计划
        
        Args:
            db: 数据库会话
            maintenance_update: 维修更新模型
            
        Returns:
            ResponseUtil: 响应结果
        """
        try:
            maintenance_data = {}
            
            # 只包含非None的字段
            if maintenance_update.scheduled_maintenance_time is not None:
                maintenance_data['scheduled_maintenance_time'] = maintenance_update.scheduled_maintenance_time
            if maintenance_update.maintenance_description is not None:
                maintenance_data['maintenance_description'] = maintenance_update.maintenance_description
            if maintenance_update.maintenance_status is not None:
                maintenance_data['maintenance_status'] = maintenance_update.maintenance_status
            if maintenance_update.maintenance_notes is not None:
                maintenance_data['maintenance_notes'] = maintenance_update.maintenance_notes
            
            success = await AlertDao.update_maintenance(
                db, 
                maintenance_update.alert_id, 
                maintenance_data,
                'system'  # 占位符，实际不使用
            )
            
            if success:
                logger.info(f"成功更新告警 {maintenance_update.alert_id} 的维修计划")
                return ResponseUtil.success(msg="更新维修计划成功")
            else:
                return ResponseUtil.failure(msg="告警不存在或更新维修计划失败")
                
        except Exception as e:
            logger.error(f"更新维修计划失败: {str(e)}")
            return ResponseUtil.failure(msg="更新维修计划失败")

    @classmethod
    async def batch_schedule_maintenance_services(
        cls,
        db: AsyncSession,
        batch_maintenance: BatchMaintenanceUpdateModel
    ) -> ResponseUtil:
        """
        批量安排维修时间
        
        Args:
            db: 数据库会话
            batch_maintenance: 批量维修模型
            
        Returns:
            ResponseUtil: 响应结果
        """
        try:
            maintenance_data = {}
            
            # 只包含非None的字段
            if batch_maintenance.scheduled_maintenance_time is not None:
                maintenance_data['scheduled_maintenance_time'] = batch_maintenance.scheduled_maintenance_time
            if batch_maintenance.maintenance_description is not None:
                maintenance_data['maintenance_description'] = batch_maintenance.maintenance_description
            if batch_maintenance.maintenance_status is not None:
                maintenance_data['maintenance_status'] = batch_maintenance.maintenance_status
            if batch_maintenance.maintenance_notes is not None:
                maintenance_data['maintenance_notes'] = batch_maintenance.maintenance_notes
            
            # 移除audit字段，直接调用DAO方法
            updated_count = await AlertDao.batch_schedule_maintenance(
                db, 
                batch_maintenance.alert_ids, 
                maintenance_data
            )
            
            if updated_count > 0:
                logger.info(f"成功批量安排维修时间，更新了 {updated_count} 条记录")
                return ResponseUtil.success(msg=f"批量安排维修时间成功，更新了 {updated_count} 条记录")
            else:
                return ResponseUtil.failure(msg="没有找到符合条件的告警记录")
                
        except Exception as e:
            logger.error(f"批量安排维修时间失败: {str(e)}")
            return ResponseUtil.failure(msg="批量安排维修时间失败")

    @classmethod
    async def get_maintenance_schedule_services(
        cls,
        db: AsyncSession,
        query_object: MaintenancePageQueryModel
    ) -> PageResponseModel:
        """
        获取维修计划列表
        
        Args:
            db: 数据库会话
            query_object: 查询对象
            
        Returns:
            PageResponseModel: 分页响应
        """
        try:
            query_dict = {}
            if query_object.device_id:
                query_dict['device_id'] = query_object.device_id
            if query_object.maintenance_status:
                query_dict['maintenance_status'] = query_object.maintenance_status
            if query_object.scheduled_start_time:
                query_dict['scheduled_start_time'] = query_object.scheduled_start_time
            if query_object.scheduled_end_time:
                query_dict['scheduled_end_time'] = query_object.scheduled_end_time
            
            maintenance_list, total = await AlertDao.get_maintenance_schedule(
                db, 
                query_dict,
                query_object.page_num,
                query_object.page_size
            )
            
            # 使用现有的分页方法创建分页响应
            has_next = math.ceil(total / query_object.page_size) > query_object.page_num if total > 0 else False
            from utils.common_util import CamelCaseUtil
            return PageResponseModel(
                rows=CamelCaseUtil.transform_result(maintenance_list),
                pageNum=query_object.page_num,
                pageSize=query_object.page_size,
                total=total,
                hasNext=has_next
            )
            
        except Exception as e:
            logger.error(f"获取维修计划列表失败: {str(e)}")
            raise

    @classmethod
    async def cancel_maintenance_services(
        cls,
        db: AsyncSession,
        alert_id: int,
        operator: str
    ) -> ResponseUtil:
        """
        取消告警维修计划
        
        Args:
            db: 数据库会话
            alert_id: 告警ID
            operator: 操作人
            
        Returns:
            ResponseUtil: 响应结果
        """
        try:
            maintenance_data = {
                'maintenance_status': 'cancelled'
            }
            
            success = await AlertDao.update_maintenance(
                db, 
                alert_id, 
                maintenance_data,
                operator
            )
            
            if success:
                logger.info(f"成功取消告警 {alert_id} 的维修计划")
                return ResponseUtil.success(msg="取消维修计划成功")
            else:
                return ResponseUtil.failure(msg="告警不存在或取消维修计划失败")
                
        except Exception as e:
            logger.error(f"取消维修计划失败: {str(e)}")
            return ResponseUtil.failure(msg="取消维修计划失败")

    @classmethod
    async def get_calendar_maintenance_services(
        cls,
        db: AsyncSession,
        start_date: str,
        end_date: str
    ):
        """
        获取日历视图的维修计划数据
        
        Args:
            db: 数据库会话
            start_date: 开始日期
            end_date: 结束日期
            
        Returns:
            CalendarMaintenanceResponseModel: 维修计划列表
        """
        try:
            from module_redfish.entity.vo.alert_vo import CalendarMaintenanceResponseModel
            # 获取指定时间范围内的维修计划
            maintenance_list = await AlertDao.get_calendar_maintenance(db, start_date, end_date)
            
            return CalendarMaintenanceResponseModel.create(maintenance_list)
            
        except Exception as e:
            logger.error(f"获取日历维修计划失败: {str(e)}")
            raise 

    @classmethod
    async def delete_alert_services(
        cls,
        db: AsyncSession,
        alert_id: int,
        delete_reason: Optional[str] = None
    ):
        """
        删除告警服务
        
        Args:
            db: 数据库会话
            alert_id: 告警ID
            delete_reason: 删除原因
            
        Returns:
            AlertDeleteResponseModel: 删除结果
        """
        try:
            from module_redfish.entity.vo.alert_vo import AlertDeleteResponseModel
            
            # 首先检查告警是否存在
            existing_alert = await AlertDao.get_alert_by_id(db, alert_id)
            if not existing_alert:
                return AlertDeleteResponseModel.create_failure("告警不存在或已被删除")
            
            # 执行删除
            success = await AlertDao.delete_alert(db, alert_id, delete_reason)
            
            if success:
                logger.info(f"告警删除成功: alert_id={alert_id}, reason={delete_reason}")
                return AlertDeleteResponseModel.create_success(1, "告警删除成功")
            else:
                return AlertDeleteResponseModel.create_failure("删除失败")
                
        except Exception as e:
            logger.error(f"删除告警服务失败: {str(e)}")
            return AlertDeleteResponseModel.create_failure("删除失败")
    
    @classmethod
    async def batch_delete_alerts_services(
        cls,
        db: AsyncSession,
        alert_ids: List[int],
        delete_reason: Optional[str] = None
    ):
        """
        批量删除告警服务
        
        Args:
            db: 数据库会话
            alert_ids: 告警ID列表
            delete_reason: 删除原因
            
        Returns:
            AlertDeleteResponseModel: 删除结果
        """
        try:
            from module_redfish.entity.vo.alert_vo import AlertDeleteResponseModel
            
            if not alert_ids:
                return AlertDeleteResponseModel.create_failure("请选择要删除的告警")
            
            # 执行批量删除
            deleted_count = await AlertDao.batch_delete_alerts(db, alert_ids, delete_reason)
            
            if deleted_count > 0:
                logger.info(f"批量删除告警成功: 删除了{deleted_count}个告警, reason={delete_reason}")
                return AlertDeleteResponseModel.create_success(deleted_count, f"成功删除{deleted_count}个告警")
            else:
                return AlertDeleteResponseModel.create_failure("没有找到可删除的告警")
                
        except Exception as e:
            logger.error(f"批量删除告警服务失败: {str(e)}")
            return AlertDeleteResponseModel.create_failure("批量删除失败") 

    @staticmethod
    async def export_alert_list_services(alert_list: List):
        """
        导出告警信息服务
        
        Args:
            alert_list: 告警信息列表
            
        Returns:
            bytes: 告警信息对应excel的二进制数据
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, PatternFill, Font
        from openpyxl.utils import get_column_letter
        import io
        
        # 创建一个映射字典，将英文键映射到中文键
        mapping_dict = {
            'serialNo': '序号',
            'hostname': '主机名',
            'businessIp': '业务IP',
            'componentType': '组件类型',
            'componentName': '组件名称',
            'urgencyLevel': '紧急程度',
            'healthStatus': '健康状态',
            'alertStatus': '告警状态',
            'firstOccurrence': '首次发生时间',
            'lastOccurrence': '最后发生时间',
            'scheduledMaintenanceTime': '计划维修时间',
            'maintenanceStatus': '维修状态',
            'maintenanceDescription': '维修描述',
            'maintenanceNotes': '维修备注',
            'createTime': '创建时间',
            'updateTime': '更新时间',
        }

        # 处理数据格式化
        processed_data = []
        for index, item in enumerate(alert_list, 1):
            # 确保 item 是字典类型，如果是对象则转换为字典
            if not isinstance(item, dict):
                # 将AlertResponseModel对象转换为字典
                if hasattr(item, 'model_dump'):
                    # 使用Pydantic的model_dump方法，自动处理驼峰命名
                    item = item.model_dump(by_alias=True)
                elif hasattr(item, '__dict__'):
                    item_dict = {}
                    for key, value in item.__dict__.items():
                        if not key.startswith('_'):  # 排除私有属性
                            # 转换为驼峰命名
                            from utils.common_util import CamelCaseUtil
                            camel_key = CamelCaseUtil.snake_to_camel(key)
                            item_dict[camel_key] = value
                    item = item_dict
                else:
                    continue
            
            # 格式化紧急程度
            urgency_level_map = {
                'urgent': '紧急',
                'scheduled': '择期'
            }
            urgency_level = item.get('urgencyLevel')
            if urgency_level:
                item['urgencyLevel'] = urgency_level_map.get(urgency_level, '未知')
            else:
                item['urgencyLevel'] = '未知'
            
            # 格式化健康状态
            health_status_map = {
                'OK': '正常',
                'ok': '正常',
                'Warning': '警告',
                'warning': '警告',
                'Critical': '警告',  # 简化分类：Critical合并到warning
                'critical': '警告',
                'Unknown': '未知',
                'unknown': '未知'
            }
            health_status = item.get('healthStatus')
            if health_status:
                item['healthStatus'] = health_status_map.get(health_status, '警告')  # 默认为警告而不是未知
            else:
                item['healthStatus'] = '警告'
            
            # 格式化告警状态
            alert_status_map = {
                'active': '活跃',
                'resolved': '已解决',
                'ignored': '已忽略'
            }
            alert_status = item.get('alertStatus')
            if alert_status:
                item['alertStatus'] = alert_status_map.get(alert_status, '未知')
            else:
                item['alertStatus'] = '未知'
            
            # 格式化维修状态
            maintenance_status_map = {
                'none': '未安排',
                'planned': '已计划',
                'in_progress': '进行中',
                'completed': '已完成',
                'cancelled': '已取消'
            }
            maintenance_status = item.get('maintenanceStatus')
            if maintenance_status:
                item['maintenanceStatus'] = maintenance_status_map.get(maintenance_status, '未安排')
            else:
                item['maintenanceStatus'] = '未安排'
            
            # 格式化时间
            time_fields = ['firstOccurrence', 'lastOccurrence', 'scheduledMaintenanceTime', 'createTime', 'updateTime']
            for time_field in time_fields:
                time_value = item.get(time_field)
                if time_value:
                    try:
                        if isinstance(time_value, str):
                            item[time_field] = time_value
                        else:
                            item[time_field] = time_value.strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        item[time_field] = ''
                else:
                    item[time_field] = ''
            
            # 添加序号
            item['serialNo'] = index
            processed_data.append(item)

        # 使用openpyxl创建Excel文件，应用和模板一样的样式
        wb = Workbook()
        ws = wb.active
        ws.title = "告警列表"
        
        # 表头样式设置（与设备导出一致）
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        # 获取表头列表
        headers = list(mapping_dict.values())
        keys = list(mapping_dict.keys())
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # 设置列宽
            ws.column_dimensions[get_column_letter(col_num)].width = 15
        
        # 写入数据
        for row_num, item in enumerate(processed_data, 2):
            for col_num, key in enumerate(keys, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = item.get(key, '')
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 保存为字节数据
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        return file_stream.getvalue() 