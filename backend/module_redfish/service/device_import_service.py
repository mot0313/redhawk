"""
设备导入导出Service层
"""
import io
import pandas as pd
from datetime import datetime
from typing import List
from fastapi import UploadFile, Request
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from module_redfish.dao.device_dao import DeviceDao
from module_redfish.entity.vo.device_vo import AddDeviceModel, EditDeviceModel
from module_redfish.core.redfish_client import encrypt_password
from module_redfish.service.business_rule_service import BusinessRuleService
from module_admin.entity.vo.user_vo import CurrentUserModel
from module_admin.entity.vo.common_vo import CrudResponseModel
from config.database import AsyncSessionLocal
from utils.log_util import logger


class DeviceImportService:
    """设备导入导出服务"""
    
    @classmethod
    async def batch_import_device_services(
        cls,
        request: Request,
        db: AsyncSession,
        file: UploadFile,
        update_support: bool,
        current_user: CurrentUserModel
    ):
        """
        批量导入设备服务
        
        Args:
            request: 请求对象
            db: 数据库会话
            file: 上传的Excel文件
            update_support: 是否支持更新已存在的设备
            current_user: 当前用户
            
        Returns:
            CrudResponseModel: 导入结果
        """
        add_error_result = []
        success_count = 0
        error_count = 0
        
        try:
            # 读取Excel文件
            df = pd.read_excel(file.file, dtype=str)
            df = df.fillna('')  # 填充空值
            
            # 验证必要的列是否存在
            required_columns = ['主机名*', '业务IP*', '带外IP*']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return CrudResponseModel(
                    is_success=False, 
                    message=f'Excel文件缺少必要的列: {", ".join(missing_columns)}'
                )
            
            # 获取业务类型数据
            business_types = []
            business_type_map = {}
            async with AsyncSessionLocal() as temp_db:
                try:
                    business_types = await BusinessRuleService.get_business_types_services(temp_db)
                    business_type_map = {bt['typeCode']: bt['typeCode'] for bt in business_types}
                except Exception as e:
                    business_types = [
                        {'typeName': '其他服务', 'description': '其他类型业务服务', 'typeCode': 'OTHER'},
                    ]
                    business_type_map = {'OTHER': 'OTHER'}
            
            for index, row in df.iterrows():
                count = index + 1
                try:
                    # 验证必填字段
                    if not row['主机名*'] or not row['业务IP*'] or not row['带外IP*']:
                        add_error_result.append(f"{count}.主机名、业务IP、带外IP为必填项")
                        error_count += 1
                        continue
                    
                    # 处理Redfish凭证
                    redfish_username = row.get('Redfish用户名', '').strip() or 'admin'
                    redfish_password = row.get('Redfish密码', '').strip() or 'admin'
                    
                    # 获取业务类型
                    business_type = row.get('业务类型', 'OTHER')
                    if business_type not in business_type_map:
                        add_error_result.append(f"{count}.业务类型无效，请使用有效的业务类型编码")
                        error_count += 1
                        continue
                    
                    # 创建设备模型
                    add_device = AddDeviceModel(
                        hostname=row['主机名*'],
                        business_ip=row['业务IP*'] if row['业务IP*'] else None,
                        oob_ip=row['带外IP*'],
                        location=row.get('机房位置', ''),
                        operating_system=row.get('操作系统', '') if row.get('操作系统', '') else None,
                        serial_number=row.get('序列号', '') if row.get('序列号', '') else None,
                        model=row.get('设备型号', '') if row.get('设备型号', '') else None,
                        manufacturer=row.get('厂商', '') if row.get('厂商', '') else None,
                        technical_system=row.get('技术系统', '') if row.get('技术系统', '') else None,
                        system_owner=row.get('系统负责人', '') if row.get('系统负责人', '') else None,
                        business_type=business_type,
                        health_status='unknown',
                        monitor_enabled=1,
                        redfish_username=redfish_username,
                        redfish_password=redfish_password,
                        remark=row.get('备注', '') if row.get('备注', '') else None,
                        create_by=current_user.user.user_name,
                        create_time=datetime.now()
                    )
                    
                    # 检查设备是否已存在
                    existing_device = await DeviceDao.get_device_by_hostname_or_ip(
                        db, add_device.hostname, add_device.oob_ip
                    )
                    
                    if existing_device:
                        if update_support:
                            # 更新已存在的设备
                            edit_device = EditDeviceModel(
                                device_id=existing_device.device_id,
                                hostname=add_device.hostname,
                                business_ip=add_device.business_ip,
                                oob_ip=add_device.oob_ip,
                                location=add_device.location,
                                operating_system=add_device.operating_system,
                                serial_number=add_device.serial_number,
                                model=add_device.model,
                                manufacturer=add_device.manufacturer,
                                technical_system=add_device.technical_system,
                                system_owner=add_device.system_owner,
                                business_type=add_device.business_type,
                                health_status=add_device.health_status,
                                monitor_enabled=add_device.monitor_enabled,
                                redfish_username=add_device.redfish_username,
                                redfish_password=add_device.redfish_password,
                                remark=add_device.remark,
                                update_by=current_user.user.user_name,
                                update_time=datetime.now()
                            )
                            
                            # 加密密码
                            edit_device.redfish_password = encrypt_password(edit_device.redfish_password)
                            
                            await DeviceDao.edit_device(db, edit_device)
                            success_count += 1
                        else:
                            add_error_result.append(f"{count}.设备 {row['主机名*']} 已存在")
                            error_count += 1
                    else:
                        # 添加新设备
                        add_device.redfish_password = encrypt_password(add_device.redfish_password)
                        await DeviceDao.add_device(db, add_device)
                        success_count += 1
                        
                except Exception as e:
                    add_error_result.append(f"{count}.导入设备 {row.get('主机名*', 'unknown')} 失败: {str(e)}")
                    error_count += 1
                    logger.error(f"导入设备失败: {str(e)}")
            
            await db.commit()
            
            result_message = f'导入完成! 成功: {success_count}, 失败: {error_count}'
            if add_error_result:
                result_message += '\n错误详情:\n' + '\n'.join(add_error_result)
                
            return CrudResponseModel(is_success=True, message=result_message)
            
        except Exception as e:
            await db.rollback()
            logger.error(f"批量导入设备失败: {str(e)}")
            return CrudResponseModel(is_success=False, message=f'导入失败: {str(e)}')

    @staticmethod
    async def get_device_import_template_services():
        """获取设备导入模板服务"""
        # 获取业务类型数据
        business_types = []
        async with AsyncSessionLocal() as db:
            try:
                business_types = await BusinessRuleService.get_business_types_services(db)
            except Exception as e:
                business_types = [
                    {'typeName': '其他服务', 'description': '其他类型业务服务'},
                ]
        
        wb = Workbook()
        ws = wb.active
        ws.title = "设备导入模板"
        
        # 定义表头
        headers = [
            '主机名*', '业务IP*', '带外IP*', '机房位置', '操作系统', '序列号', 
            '设备型号', '厂商', '技术系统', '系统负责人', '业务类型', 
            'Redfish用户名', 'Redfish密码', '备注'
        ]
        
        # 示例数据
        example_data = [
            [
                'server-001', '192.168.1.101', '192.168.100.101', 'XW_B1B07_25-26', 
                'CentOS 7.9', 'SN20231001', 'PowerEdge R750', 'Dell', 
                '虚拟化平台', '张三', 'OB-GMT-3N', 'admin', 'admin', '生产环境核心服务器'
            ],
            [
                'server-002', '192.168.1.102', '192.168.100.102', 'YJ_F3D13_24-25', 
                'Ubuntu 20.04', 'SN20231002', 'ProLiant DL380', 'HPE', 
                '数据库服务', '李四', 'OB-GMT-3N', 'root', 'password', '数据库集群主节点'
            ],
            [
                'server-003', '192.168.1.103', '192.168.100.103', 'YZ_B6C10_14-15', 
                'Ubuntu 20.04', 'SN20231003', 'ProLiant DL380', 'HPE', 
                '测试', '王五', 'OB-DATA-5N', 'Administrator', 'Password', '测试环境核心服务器'
            ]
        ]
        
        # 设置表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(col_num)].width = 15
        
        # 写入示例数据
        for row_num, data_row in enumerate(example_data, 2):
            for col_num, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加说明工作表
        ws_info = wb.create_sheet("导入说明")
        
        info_content = [
            ["设备导入模板使用说明", ""],
            ["", ""],
            ["必填字段（带*号）:", ""],
            ["主机名*", "设备的主机名，必须唯一"],
            ["业务IP*", "设备的业务网络IP地址"],
            ["带外IP*", "设备的BMC管理IP地址，必须唯一"],
            ["", ""],
            ["可选字段:", ""],
            ["机房位置", "设备所在的物理位置"],
            ["操作系统", "设备运行的操作系统"],
            ["序列号", "设备的序列号"],
            ["设备型号", "设备的具体型号"],
            ["厂商", "设备制造商"],
            ["技术系统", "设备所属的技术系统"],
            ["系统负责人", "设备的负责人姓名"],
            ["业务类型", "设备承载的业务类型"],
            ["Redfish用户名", "BMC登录用户名（默认：admin）"],
            ["Redfish密码", "BMC登录密码（默认：admin）"],
            ["备注", "设备的备注信息"],
            ["", ""],
            ["业务类型选项:", ""]
        ]
        
        # 添加业务类型
        for bt in business_types:
            type_code = bt.get('typeCode', '')
            type_name = bt.get('typeName', '')
            info_content.append([type_code, type_name])
        
        info_content.extend([
            ["", ""],
            ["注意事项:", ""],
            ["1. 带*号的字段为必填项，不能为空"],
            ["2. Redfish用户名和密码如果不填写，系统将默认设置为admin"],
            ["3. 设备的健康状态和监控状态由系统自动设置"],
            ["4. 导入时会检查主机名和带外IP的唯一性"],
            ["5. 支持更新已存在的设备数据"],
        ])
        
        # 写入说明内容
        for row_num, content in enumerate(info_content, 1):
            if len(content) == 2:
                key, value = content
                ws_info.cell(row=row_num, column=1, value=key)
                ws_info.cell(row=row_num, column=2, value=value)
            else:
                ws_info.cell(row=row_num, column=1, value=content[0])
            
        # 设置说明页样式
        ws_info.column_dimensions['A'].width = 25
        ws_info.column_dimensions['B'].width = 50
        
        title_cell = ws_info.cell(row=1, column=1)
        title_cell.font = Font(size=16, bold=True, color='4472C4')
        
        # 保存为字节数据
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        return file_stream.getvalue()

    @staticmethod
    async def export_device_list_services(device_list: List):
        """
        导出设备信息服务
        
        Args:
            device_list: 设备信息列表
            
        Returns:
            bytes: 设备信息Excel的二进制数据
        """
        # 创建映射字典
        mapping_dict = {
            'serialNo': '序号',
            'hostname': '主机名',
            'businessIp': '业务IP',
            'oobIp': '带外IP',
            'location': '机房位置',
            'operatingSystem': '操作系统',
            'serialNumber': '序列号',
            'model': '设备型号',
            'manufacturer': '厂商',
            'technicalSystem': '技术系统',
            'systemOwner': '系统负责人',
            'businessType': '业务类型',
            'redfishUsername': 'Redfish用户名',
            'healthStatus': '健康状态',
            'monitorEnabled': '监控状态',
            'lastCheckTime': '最后检查时间',
            'createBy': '创建者',
            'createTime': '创建时间',
            'updateBy': '更新者',
            'updateTime': '更新时间',
            'remark': '备注',
        }

        # 处理数据格式化
        processed_data = []
        for index, item in enumerate(device_list, 1):
            # 确保item是字典类型
            if not isinstance(item, dict):
                if hasattr(item, 'model_dump'):
                    item = item.model_dump(by_alias=True)
                elif hasattr(item, '__dict__'):
                    item_dict = {}
                    for key, value in item.__dict__.items():
                        if not key.startswith('_'):
                            from utils.common_util import CamelCaseUtil
                            camel_key = CamelCaseUtil.snake_to_camel(key)
                            item_dict[camel_key] = value
                    item = item_dict
                else:
                    continue
            
            # 格式化健康状态
            health_status_map = {
                'ok': '正常',
                'warning': '警告',
                'critical': '警告',
                'unknown': '未知'
            }
            health_status = item.get('healthStatus')
            if health_status:
                item['healthStatus'] = health_status_map.get(health_status, '未知')
            else:
                item['healthStatus'] = '未知'
            
            # 格式化监控状态
            monitor_enabled = item.get('monitorEnabled')
            if monitor_enabled == 1:
                item['monitorEnabled'] = '启用'
            else:
                item['monitorEnabled'] = '停用'
                
            # 隐藏密码信息
            if 'redfishPassword' in item:
                del item['redfishPassword']
            
            # 格式化时间
            time_fields = ['lastCheckTime', 'createTime', 'updateTime']
            for time_field in time_fields:
                time_value = item.get(time_field)
                if time_value:
                    try:
                        if isinstance(time_value, str):
                            item[time_field] = time_value
                        else:
                            item[time_field] = time_value.strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        item[time_field] = ''
            
            # 添加序号
            item['serialNo'] = index
            processed_data.append(item)

        # 创建Excel文件
        wb = Workbook()
        ws = wb.active
        ws.title = "设备列表"
        
        # 表头样式设置
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        # 获取表头列表
        headers = list(mapping_dict.values())
        keys = list(mapping_dict.keys())
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(col_num)].width = 15
        
        # 写入数据
        for row_num, item in enumerate(processed_data, 2):
            for col_num, key in enumerate(keys, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = item.get(key, '')
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 保存为字节数据
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        return file_stream.getvalue()
