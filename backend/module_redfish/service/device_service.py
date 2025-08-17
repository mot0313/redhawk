"""
设备管理Service层
"""
import math
import pandas as pd
from fastapi import UploadFile, Request
from datetime import datetime
from typing import List, Optional, Dict, Any
from sqlalchemy.ext.asyncio import AsyncSession
from module_redfish.dao.device_dao import DeviceDao
from module_redfish.dao.alert_dao import AlertDao
from module_redfish.entity.do.device_do import DeviceInfoDO
from module_redfish.entity.vo.device_vo import (
    DevicePageQueryModel, AddDeviceModel, EditDeviceModel, DeleteDeviceModel,
    DeviceDetailModel, DeviceHealthModel, DeviceTestConnectionModel, DeviceConnectionResult,
    DeviceModel
)
from module_redfish.redfish_client import RedfishClient, encrypt_password, decrypt_password
from module_admin.entity.vo.user_vo import CurrentUserModel
from module_admin.entity.vo.common_vo import CrudResponseModel
from utils.response_util import ResponseUtil
from utils.page_util import PageResponseModel, PageUtil
from utils.log_util import logger
from utils.excel_util import ExcelUtil
from module_redfish.service.business_rule_service import BusinessRuleService
from module_redfish.celery_tasks import recalculate_urgency_for_device
from config.database import AsyncSessionLocal
import asyncio
import subprocess
import socket
import platform


class DeviceService:
    """设备管理服务"""
    
    @classmethod
    async def get_device_list_services(
        cls,
        db: AsyncSession,
        query_object: DevicePageQueryModel,
        is_page: bool = False
    ) -> PageResponseModel:
        """
        获取设备列表
        
        Args:
            db: 数据库会话
            query_object: 查询对象
            is_page: 是否分页
            
        Returns:
            PageResponseModel: 分页响应
        """
        device_list, total = await DeviceDao.get_device_list(db, query_object, is_page)
        
        # 获取业务类型映射字典（type_code -> type_name）
        business_type_mapping = {}
        try:
            business_types = await BusinessRuleService.get_business_types_services(db)
            business_type_mapping = {bt['typeCode']: bt['typeName'] for bt in business_types}
        except Exception as e:
            logger.warning(f"获取业务类型映射失败: {str(e)}")
        
        # 解密密码字段并转换业务类型
        for device in device_list:
            device.redfish_password = "******"  # 隐藏密码
            # 将 type_code 转换为 type_name
            if device.business_type and device.business_type in business_type_mapping:
                device.business_type = business_type_mapping[device.business_type]
        
        if is_page:
            # 使用DevicePageResponseModel创建分页响应
            from module_redfish.entity.vo.device_vo import DevicePageResponseModel
            # 先转换为字典列表
            device_dict_list = [device.__dict__.copy() for device in device_list]
            for device_dict in device_dict_list:
                device_dict.pop('_sa_instance_state', None)
            return DevicePageResponseModel.create(
                devices=device_dict_list,
                page_num=query_object.page_num,
                page_size=query_object.page_size,
                total=total
            )
        else:
            # 非分页模式，返回设备列表
            from module_redfish.entity.vo.device_vo import DeviceResponseModel
            device_dict_list = [device.__dict__.copy() for device in device_list]
            for device_dict in device_dict_list:
                device_dict.pop('_sa_instance_state', None)
            return [DeviceResponseModel(**device_dict) for device_dict in device_dict_list]
    
    @classmethod
    async def get_device_detail_services(cls, db: AsyncSession, device_id: int, for_edit: bool = False) -> DeviceDetailModel:
        """
        获取设备详情
        
        Args:
            db: 数据库会话
            device_id: 设备ID
            
        Returns:
            DeviceDetailModel: 设备详情
        """
        device = await DeviceDao.get_device_by_id(db, device_id)
        if not device:
            raise ValueError("设备不存在")
        
        # 根据用途决定业务类型显示格式
        business_type_display = device.business_type
        if not for_edit:
            # 用于显示时，将 type_code 转换为 type_name
            try:
                business_types = await BusinessRuleService.get_business_types_services(db)
                business_type_mapping = {bt['typeCode']: bt['typeName'] for bt in business_types}
                if device.business_type and device.business_type in business_type_mapping:
                    business_type_display = business_type_mapping[device.business_type]
            except Exception as e:
                logger.warning(f"获取业务类型映射失败: {str(e)}")
        # 用于编辑时，保持原始的 type_code
        
        # 将DeviceInfo转换为DeviceModel
        # 使用字典方式创建，让Pydantic自动处理字段名转换
        device_dict = {
            'device_id': device.device_id,
            'hostname': device.hostname,
            'business_ip': device.business_ip,
            'oob_ip': device.oob_ip,
            'oob_port': device.oob_port,
            'location': device.location,
            'operating_system': device.operating_system,
            'serial_number': device.serial_number,
            'model': device.model,
            'manufacturer': device.manufacturer,
            'technical_system': device.technical_system,
            'system_owner': device.system_owner,
            'business_type': business_type_display,  # 使用转换后的业务类型名称
            'redfish_username': device.redfish_username,
            'redfish_password': "******",  # 隐藏密码
            'monitor_enabled': device.monitor_enabled,
            'last_check_time': device.last_check_time,
            'health_status': device.health_status,
            'create_by': device.create_by,
            'create_time': device.create_time,
            'update_by': device.update_by,
            'update_time': device.update_time,
            'remark': device.remark
        }
        
        device_model = DeviceModel.model_validate(device_dict)
        
        # 获取设备健康状态（从告警表或缓存中获取）
        # TODO: 实现健康状态获取逻辑
        health_status = device.health_status if device.health_status else "Unknown"
        last_check_time = device.last_check_time
        connection_status = "Unknown"
        
        return DeviceDetailModel(
            device=device_model,
            health_status=health_status,
            last_check_time=last_check_time,
            connection_status=connection_status
        )
    
    @classmethod
    async def add_device_services(cls, db: AsyncSession, device: AddDeviceModel) -> ResponseUtil:
        """
        添加设备
        
        Args:
            db: 数据库会话
            device: 设备信息
            
        Returns:
            ResponseUtil: 响应结果
        """
        # 检查IP是否已存在
        existing_device = await DeviceDao.get_device_by_ip(
            db, device.business_ip, device.oob_ip
        )
        if existing_device:
            return ResponseUtil.failure(msg="业务IP或带外IP已存在")
        
        # 仅在提供了密码时加密
        if device.redfish_password:
            device.redfish_password = encrypt_password(device.redfish_password)
        
        try:
            new_device = await DeviceDao.add_device(db, device)
            logger.info(f"成功添加设备: {device.hostname}")
            return ResponseUtil.success(msg="添加设备成功")
        except Exception as e:
            logger.error(f"添加设备失败: {str(e)}")
            return ResponseUtil.failure(msg="添加设备失败")
    
    @classmethod
    async def edit_device_services(cls, db: AsyncSession, device: EditDeviceModel) -> ResponseUtil:
        """
        编辑设备
        
        Args:
            db: 数据库会话
            device: 设备信息
            
        Returns:
            ResponseUtil: 响应结果
        """
        # 检查设备是否存在
        existing_device = await DeviceDao.get_device_by_id(db, device.device_id)
        if not existing_device:
            return ResponseUtil.failure(msg="设备不存在")
        
        # 记录旧的业务类型
        old_business_type = existing_device.business_type
        
        # 检查IP是否与其他设备冲突
        ip_conflict_device = await DeviceDao.get_device_by_ip(
            db, device.business_ip, device.oob_ip
        )
        if ip_conflict_device and ip_conflict_device.device_id != device.device_id:
            return ResponseUtil.failure(msg="业务IP或带外IP与其他设备冲突")
        
        # 如果密码不是掩码，则加密
        if device.redfish_password and device.redfish_password != "******":
            device.redfish_password = encrypt_password(device.redfish_password)
        else:
            # 保持原密码不变
            device.redfish_password = existing_device.redfish_password
        
        try:
            success = await DeviceDao.edit_device(db, device)
            if success:
                logger.info(f"成功编辑设备: {device.hostname}")
                
                # 检查业务类型是否变更，如果变更则触发后台任务
                if device.business_type and old_business_type != device.business_type:
                    logger.info(f"业务类型变更，为设备 {device.device_id} 触发紧急度重新计算任务。")
                    recalculate_urgency_for_device.delay(device.device_id)

                return ResponseUtil.success(msg="编辑设备成功")
            else:
                return ResponseUtil.failure(msg="编辑设备失败")
        except Exception as e:
            logger.error(f"编辑设备失败: {str(e)}")
            return ResponseUtil.failure(msg="编辑设备失败")
    
    @classmethod
    async def delete_device_services(cls, db: AsyncSession, delete_device: DeleteDeviceModel) -> ResponseUtil:
        """
        删除设备
        
        Args:
            db: 数据库会话
            delete_device: 删除设备模型
            
        Returns:
            ResponseUtil: 响应结果
        """
        device_ids_list = [int(id_str) for id_str in delete_device.device_ids.split(',')]
        if not device_ids_list:
            return ResponseUtil.failure(msg="没有提供有效的设备ID")
        
        try:
            # 1. 逻辑删除关联的告警（保留告警数据用于统计）
            deleted_alerts_count = await AlertDao.logical_delete_alerts_by_device_ids(db, device_ids_list)
            logger.info(f"逻辑删除了 {deleted_alerts_count} 条与待删除设备相关的告警。")
            
            # 提交告警逻辑删除的事务
            await db.commit()
            
            # 2. 物理删除设备
            deleted_devices_count = await DeviceDao.delete_device_by_ids(db, delete_device.device_ids)
            
            if deleted_devices_count > 0:
                logger.info(f"成功物理删除了 {deleted_devices_count} 个设备。")
                return ResponseUtil.success(msg=f"成功删除 {deleted_devices_count} 个设备，并标记删除了 {deleted_alerts_count} 条相关告警（告警数据已保留用于统计）。")
            else:
                await db.rollback()
                return ResponseUtil.failure(msg="没有设备被删除。")
                
        except Exception as e:
            await db.rollback()
            logger.error(f"删除设备时发生错误: {str(e)}")
            return ResponseUtil.failure(msg=f"删除设备时发生错误: {e}")
    

    
    @classmethod
    async def test_device_connection_by_id_services(
        cls, 
        db: AsyncSession,
        device_id: int
    ) -> DeviceConnectionResult:
        """
        通过设备ID测试设备连接
        
        Args:
            db: 数据库会话
            device_id: 设备ID
            
        Returns:
            DeviceConnectionResult: 连接测试结果
        """
        try:
            # 从数据库获取设备信息
            device = await DeviceDao.get_device_by_id(db, device_id)
            if not device:
                return DeviceConnectionResult(
                    success=False,
                    message="设备不存在",
                    system_info=None
                )
            
            # 解密密码
            decrypted_password = decrypt_password(device.redfish_password)
            
            # 创建Redfish客户端（测试连接使用更短超时，避免前端10s axios超时）
            client = RedfishClient(
                host=device.oob_ip,
                username=device.redfish_username,
                password=decrypted_password,
                port=device.oob_port,
                timeout=6
            )
            
            # 尝试连接（再加一层总超时保护）
            try:
                success = await asyncio.wait_for(client.connect(), timeout=6)
            except asyncio.TimeoutError:
                logger.warning("Redfish connect timeout (8s) during test_device_connection_by_id")
                return DeviceConnectionResult(
                    success=False,
                    message="连接测试失败",
                    system_info=None
                )
            
            if success:
                # 获取系统信息（同样增加总超时保护）
                try:
                    system_info = await asyncio.wait_for(client.get_system_info(), timeout=6)
                except asyncio.TimeoutError:
                    logger.warning("Redfish get_system_info timeout (8s) during test_device_connection_by_id")
                    await client.disconnect()
                    return DeviceConnectionResult(
                        success=False,
                        message="连接测试失败",
                        system_info=None
                    )
                await client.disconnect()
                
                return DeviceConnectionResult(
                    success=True,
                    message="连接成功",
                    system_info=system_info
                )
            else:
                return DeviceConnectionResult(
                    success=False,
                    message="连接测试失败",
                    system_info=None
                )
        except Exception as e:
            logger.error(f"测试设备连接失败 (设备ID: {device_id}): {str(e)}")
            return DeviceConnectionResult(
                success=False,
                message="连接测试失败",
                system_info=None
            )
    
    @classmethod
    async def get_monitoring_devices_services(cls, db: AsyncSession) -> List[Dict[str, Any]]:
        """
        获取启用监控的设备列表
        
        Args:
            db: 数据库会话
            
        Returns:
            List[Dict[str, Any]]: 设备列表
        """
        devices = await DeviceDao.get_monitoring_devices(db)
        
        result = []
        for device in devices:
            # 解密密码用于监控
            decrypted_password = decrypt_password(device.redfish_password)
            
            result.append({
                'device_id': device.device_id,
                'hostname': device.hostname,
                'business_ip': device.business_ip,
                'oob_ip': device.oob_ip,
                'oob_port': device.oob_port,
                'redfish_username': device.redfish_username,
                'redfish_password': decrypted_password,
                'location': device.location,
                'technical_system': device.technical_system,
                'manufacturer': device.manufacturer,
                'model': device.model
            })
        
        return result
    
    @classmethod
    async def update_device_system_info_services(
        cls,
        db: AsyncSession,
        device_id: int,
        system_info: Dict[str, Any]
    ) -> bool:
        """
        更新设备系统信息
        
        Args:
            db: 数据库会话
            device_id: 设备ID
            system_info: 系统信息
            
        Returns:
            bool: 是否成功
        """
        try:
            success = await DeviceDao.update_device_system_info(db, device_id, system_info)
            if success:
                logger.info(f"成功更新设备 {device_id} 的系统信息")
            return success
        except Exception as e:
            logger.error(f"更新设备系统信息失败: {str(e)}")
            return False
    
    @classmethod
    async def get_device_statistics_services(cls, db: AsyncSession):
        """
        获取设备统计信息
        
        Args:
            db: 数据库会话
            
        Returns:
            DeviceStatsResponseModel: 统计信息
        """
        from module_redfish.entity.vo.device_vo import DeviceStatsResponseModel
        stats = await DeviceDao.get_device_statistics(db)
        return DeviceStatsResponseModel.create(stats)

    @classmethod
    async def batch_import_device_services(
        cls,
        request: Request,
        db: AsyncSession,
        file: UploadFile,
        update_support: bool,
        current_user: CurrentUserModel
    ):
        """
        批量导入设备服务
        
        Args:
            request: 请求对象
            db: 数据库会话
            file: 上传的Excel文件
            update_support: 是否支持更新已存在的设备
            current_user: 当前用户
            
        Returns:
            CrudResponseModel: 导入结果
        """
        add_error_result = []
        success_count = 0
        error_count = 0
        
        try:
            # 读取Excel文件
            df = pd.read_excel(file.file, dtype=str)
            df = df.fillna('')  # 填充空值
            
            # 验证必要的列是否存在
            required_columns = ['主机名*', '业务IP*', '带外IP*']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return CrudResponseModel(
                    is_success=False, 
                    message=f'Excel文件缺少必要的列: {", ".join(missing_columns)}'
                )
            
            # 获取业务类型数据
            business_types = []
            business_type_map = {}  # 用于存储code到验证的映射
            async with AsyncSessionLocal() as db:
                try:
                    business_types = await BusinessRuleService.get_business_types_services(db)
                    # 创建映射表，支持通过code验证
                    business_type_map = {bt['typeCode']: bt['typeCode'] for bt in business_types}
                except Exception as e:
                    # 如果获取失败，使用默认值
                    business_types = [
                        {'typeName': '其他服务', 'description': '其他类型业务服务', 'typeCode': 'OTHER'},
                    ]
                    business_type_map = {'OTHER': 'OTHER'}
            
            for index, row in df.iterrows():
                count = index + 1
                try:
                    # 验证必填字段
                    if not row['主机名*'] or not row['业务IP*'] or not row['带外IP*']:
                        add_error_result.append(f"{count}.主机名、业务IP、带外IP为必填项")
                        error_count += 1
                        continue
                    
                    # 处理Redfish凭证，如果不填写则默认为admin
                    redfish_username = row.get('Redfish用户名', '').strip() or 'admin'
                    redfish_password = row.get('Redfish密码', '').strip() or 'admin'
                    
                    # 获取业务类型
                    business_type = row.get('业务类型', 'OTHER')
                    if business_type not in business_type_map:
                        add_error_result.append(f"{count}.业务类型无效，请使用有效的业务类型编码")
                        error_count += 1
                        continue
                    
                    # 创建设备模型
                    add_device = AddDeviceModel(
                        hostname=row['主机名*'],
                        business_ip=row['业务IP*'] if row['业务IP*'] else None,
                        oob_ip=row['带外IP*'],
                        location=row.get('机房位置', ''),
                        operating_system=row.get('操作系统', '') if row.get('操作系统', '') else None,
                        serial_number=row.get('序列号', '') if row.get('序列号', '') else None,
                        model=row.get('设备型号', '') if row.get('设备型号', '') else None,
                        manufacturer=row.get('厂商', '') if row.get('厂商', '') else None,
                        technical_system=row.get('技术系统', '') if row.get('技术系统', '') else None,
                        system_owner=row.get('系统负责人', '') if row.get('系统负责人', '') else None,
                        business_type=business_type,
                        health_status='unknown',  # 默认为unknown
                        monitor_enabled=1,  # 默认启用监控
                        redfish_username=redfish_username,
                        redfish_password=redfish_password,
                        remark=row.get('备注', '') if row.get('备注', '') else None,
                        create_by=current_user.user.user_name,
                        create_time=datetime.now()
                    )
                    
                    # 检查设备是否已存在（通过主机名或带外IP）
                    existing_device = await DeviceDao.get_device_by_hostname_or_ip(
                        db, add_device.hostname, add_device.oob_ip
                    )
                    
                    if existing_device:
                        if update_support:
                            # 更新已存在的设备
                            edit_device = EditDeviceModel(
                                device_id=existing_device.device_id,
                                hostname=add_device.hostname,
                                business_ip=add_device.business_ip,
                                oob_ip=add_device.oob_ip,
                                location=add_device.location,
                                operating_system=add_device.operating_system,
                                serial_number=add_device.serial_number,
                                model=add_device.model,
                                manufacturer=add_device.manufacturer,
                                technical_system=add_device.technical_system,
                                system_owner=add_device.system_owner,
                                business_type=add_device.business_type,
                                health_status=add_device.health_status,
                                monitor_enabled=add_device.monitor_enabled,
                                redfish_username=add_device.redfish_username,
                                redfish_password=add_device.redfish_password,
                                remark=add_device.remark,
                                update_by=current_user.user.user_name,
                                update_time=datetime.now()
                            )
                            
                            # 加密密码
                            edit_device.redfish_password = encrypt_password(edit_device.redfish_password)
                            
                            await DeviceDao.edit_device(db, edit_device)
                            success_count += 1
                        else:
                            add_error_result.append(f"{count}.设备 {row['主机名*']} 已存在")
                            error_count += 1
                    else:
                        # 添加新设备
                        # 加密密码
                        add_device.redfish_password = encrypt_password(add_device.redfish_password)
                        
                        await DeviceDao.add_device(db, add_device)
                        success_count += 1
                        
                except Exception as e:
                    add_error_result.append(f"{count}.导入设备 {row.get('主机名*', 'unknown')} 失败: {str(e)}")
                    error_count += 1
                    logger.error(f"导入设备失败: {str(e)}")
            
            await db.commit()
            
            result_message = f'导入完成! 成功: {success_count}, 失败: {error_count}'
            if add_error_result:
                result_message += '\n错误详情:\n' + '\n'.join(add_error_result)
                
            return CrudResponseModel(is_success=True, message=result_message)
            
        except Exception as e:
            await db.rollback()
            logger.error(f"批量导入设备失败: {str(e)}")
            return CrudResponseModel(is_success=False, message=f'导入失败: {str(e)}')

    @staticmethod
    async def get_device_import_template_services():
        """
        获取设备导入模板服务
        
        Returns:
            bytes: 设备导入模板Excel的二进制数据
        """
        # 创建Excel工作簿
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, PatternFill, Font
        from openpyxl.utils import get_column_letter
        import io
        
        # 获取业务类型数据
        business_types = []
        async with AsyncSessionLocal() as db:
            try:
                business_types = await BusinessRuleService.get_business_types_services(db)
            except Exception as e:
                # 如果获取失败，使用默认值
                business_types = [
                    {'typeName': '其他服务', 'description': '其他类型业务服务'},
                ]
        
        wb = Workbook()
        ws = wb.active
        ws.title = "设备导入模板"
        
        # 定义表头
        headers = [
            '主机名*', '业务IP*', '带外IP*', '机房位置', '操作系统', '序列号', 
            '设备型号', '厂商', '技术系统', '系统负责人', '业务类型', 
            'Redfish用户名', 'Redfish密码', '备注'
        ]
        
        # 示例数据
        example_data = [
            [
                'server-001', '192.168.1.101', '192.168.100.101', 'XW_B1B07_25-26', 
                'CentOS 7.9', 'SN20231001', 'PowerEdge R750', 'Dell', 
                '虚拟化平台', '张三', 'OB-GMT-3N', 'admin', 'admin', '生产环境核心服务器'
            ],
            [
                'server-002', '192.168.1.102', '192.168.100.102', 'YJ_F3D13_24-25', 
                'Ubuntu 20.04', 'SN20231002', 'ProLiant DL380', 'HPE', 
                '数据库服务', '李四', 'OB-GMT-3N', 'root', 'password', '数据库集群主节点'
            ],
            [
                'server-003', '192.168.1.103', '192.168.100.103', 'YZ_B6C10_14-15', 
                'Ubuntu 20.04', 'SN20231003', 'ProLiant DL380', 'HPE', 
                '测试', '王五', 'OB-DATA-5N', 'Administrator', 'Password', '测试环境核心服务器'
            ]
        ]
        
        # 设置表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # 设置列宽
            ws.column_dimensions[get_column_letter(col_num)].width = 15
        
        # 写入示例数据
        for row_num, data_row in enumerate(example_data, 2):
            for col_num, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加说明工作表
        ws_info = wb.create_sheet("导入说明")
        
        # 说明内容
        info_content = [
            ["设备导入模板使用说明", ""],
            ["", ""],
            ["必填字段（带*号）:", ""],
            ["主机名*", "设备的主机名，必须唯一"],
            ["业务IP*", "设备的业务网络IP地址"],
            ["带外IP*", "设备的BMC管理IP地址，必须唯一"],
            ["", ""],
            ["可选字段:", ""],
            ["机房位置", "设备所在的物理位置"],
            ["操作系统", "设备运行的操作系统"],
            ["序列号", "设备的序列号"],
            ["设备型号", "设备的具体型号"],
            ["厂商", "设备制造商"],
            ["技术系统", "设备所属的技术系统"],
            ["系统负责人", "设备的负责人姓名"],
            ["业务类型", "设备承载的业务类型"],
            ["Redfish用户名", "BMC登录用户名（默认：admin）"],
            ["Redfish密码", "BMC登录密码（默认：admin）"],
            ["备注", "设备的备注信息"],
            ["", ""],
            ["业务类型选项:", ""]
        ]
        
        # 添加从数据库获取的业务类型
        for bt in business_types:
            type_code = bt.get('typeCode', '')
            type_name = bt.get('typeName', '')
            description = bt.get('description', '')
            # 左侧显示type_code，右侧显示type_name
            info_content.append([type_code, type_name])
        
        # 添加剩余说明
        info_content.extend([
            ["", ""],
            ["注意事项:", ""],
            ["1. 带*号的字段为必填项，不能为空"],
            ["2. Redfish用户名和密码如果不填写，系统将默认设置为admin"],
            ["3. 设备的健康状态和监控状态由系统自动设置"],
            ["4. 导入时会检查主机名和带外IP的唯一性"],
            ["5. 支持更新已存在的设备数据"],
        ])
        
        # 写入说明内容
        for row_num, content in enumerate(info_content, 1):
            if len(content) == 2:
                key, value = content
                ws_info.cell(row=row_num, column=1, value=key)
                ws_info.cell(row=row_num, column=2, value=value)
            else:
                # 只有一个值的情况
                ws_info.cell(row=row_num, column=1, value=content[0])
            
        # 设置说明页样式
        ws_info.column_dimensions['A'].width = 25
        ws_info.column_dimensions['B'].width = 50
        
        # 设置标题样式
        title_cell = ws_info.cell(row=1, column=1)
        title_cell.font = Font(size=16, bold=True, color='4472C4')
        
        # 保存为字节数据
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        return file_stream.getvalue()

    @staticmethod
    async def export_device_list_services(device_list: List):
        """
        导出设备信息服务
        
        Args:
            device_list: 设备信息列表
            
        Returns:
            bytes: 设备信息对应excel的二进制数据
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, PatternFill, Font
        from openpyxl.utils import get_column_letter
        import io
        
        # 创建一个映射字典，将英文键映射到中文键
        mapping_dict = {
            'serialNo': '序号',
            'hostname': '主机名',
            'businessIp': '业务IP',
            'oobIp': '带外IP',
            'location': '机房位置',
            'operatingSystem': '操作系统',
            'serialNumber': '序列号',
            'model': '设备型号',
            'manufacturer': '厂商',
            'technicalSystem': '技术系统',
            'systemOwner': '系统负责人',
            'businessType': '业务类型',
            'redfishUsername': 'Redfish用户名',
            'healthStatus': '健康状态',
            'monitorEnabled': '监控状态',
            'lastCheckTime': '最后检查时间',
            'createBy': '创建者',
            'createTime': '创建时间',
            'updateBy': '更新者',
            'updateTime': '更新时间',
            'remark': '备注',
        }

        # 处理数据格式化
        processed_data = []
        for index, item in enumerate(device_list, 1):
            # 确保 item 是字典类型，如果是对象则转换为字典
            if not isinstance(item, dict):
                # 将DeviceResponseModel对象转换为字典
                if hasattr(item, 'model_dump'):
                    # 使用Pydantic的model_dump方法，自动处理驼峰命名
                    item = item.model_dump(by_alias=True)
                elif hasattr(item, '__dict__'):
                    item_dict = {}
                    for key, value in item.__dict__.items():
                        if not key.startswith('_'):  # 排除私有属性
                            # 转换为驼峰命名
                            from utils.common_util import CamelCaseUtil
                            camel_key = CamelCaseUtil.snake_to_camel(key)
                            item_dict[camel_key] = value
                    item = item_dict
                else:
                    continue
            
            # 格式化健康状态
            health_status_map = {
                'ok': '正常',
                'warning': '警告',
                'critical': '警告',  # 简化分类：critical合并到warning
                'unknown': '未知'
            }
            health_status = item.get('healthStatus')
            if health_status:
                item['healthStatus'] = health_status_map.get(health_status, '未知')
            else:
                item['healthStatus'] = '未知'
            
            # 格式化监控状态
            monitor_enabled = item.get('monitorEnabled')
            if monitor_enabled == 1:
                item['monitorEnabled'] = '启用'
            else:
                item['monitorEnabled'] = '停用'
                
            # 隐藏密码信息（不导出）
            if 'redfishPassword' in item:
                del item['redfishPassword']
            
            # 格式化时间
            time_fields = ['lastCheckTime', 'createTime', 'updateTime']
            for time_field in time_fields:
                time_value = item.get(time_field)
                if time_value:
                    try:
                        if isinstance(time_value, str):
                            item[time_field] = time_value
                        else:
                            item[time_field] = time_value.strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        item[time_field] = ''
            
            # 添加序号
            item['serialNo'] = index
            processed_data.append(item)

        # 使用openpyxl创建Excel文件，应用和模板一样的样式
        wb = Workbook()
        ws = wb.active
        ws.title = "设备列表"
        
        # 表头样式设置（与模板一致）
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        # 获取表头列表
        headers = list(mapping_dict.values())
        keys = list(mapping_dict.keys())
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # 设置列宽
            ws.column_dimensions[get_column_letter(col_num)].width = 15
        
        # 写入数据
        for row_num, item in enumerate(processed_data, 2):
            for col_num, key in enumerate(keys, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = item.get(key, '')
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 保存为字节数据
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        return file_stream.getvalue()

    @classmethod
    async def check_business_ip_connectivity_services(
        cls,
        db: AsyncSession,
        device_id: int = None,
        business_ip: str = None
    ) -> Dict[str, Any]:
        """
        检测设备业务IP连通性（基于ping和常用端口）
        
        Args:
            db: 数据库会话
            device_id: 设备ID（二选一）
            business_ip: 业务IP（二选一）
            
        Returns:
            Dict[str, Any]: 连通性检测结果
        """
        if device_id:
            device = await DeviceDao.get_device_by_id(db, device_id)
            if not device:
                return {"online": False, "error": "设备不存在", "check_time": datetime.now().isoformat()}
            target_ip = device.business_ip
            hostname = device.hostname
        elif business_ip:
            target_ip = business_ip
            hostname = business_ip
        else:
            return {"online": False, "error": "必须提供device_id或business_ip", "check_time": datetime.now().isoformat()}
        
        if not target_ip:
            return {"online": False, "error": "设备业务IP为空", "check_time": datetime.now().isoformat()}
        
        # 并发执行多种检测方法
        ping_task = asyncio.create_task(cls._ping_check(target_ip))
        ssh_task = asyncio.create_task(cls._tcp_port_check(target_ip, 22, timeout=2))
        http_task = asyncio.create_task(cls._tcp_port_check(target_ip, 80, timeout=2))
        https_task = asyncio.create_task(cls._tcp_port_check(target_ip, 443, timeout=2))
        
        # 等待所有检测完成
        ping_result, ssh_result, http_result, https_result = await asyncio.gather(
            ping_task, ssh_task, http_task, https_task, return_exceptions=True
        )
        
        # 处理异常结果
        def safe_result(result, method):
            if isinstance(result, Exception):
                return {"success": False, "method": method, "error": str(result)}
            return result
        
        ping_result = safe_result(ping_result, "ping")
        ssh_result = safe_result(ssh_result, "tcp_port_22")
        http_result = safe_result(http_result, "tcp_port_80") 
        https_result = safe_result(https_result, "tcp_port_443")
        
        # 综合判断：ping通或任一端口可达即认为在线
        online = (ping_result.get("success", False) or 
                 ssh_result.get("success", False) or 
                 http_result.get("success", False) or 
                 https_result.get("success", False))
        
        return {
            "online": online,
            "hostname": hostname,
            "business_ip": target_ip,
            "ping": ping_result,
            "ssh_port": ssh_result,
            "http_port": http_result,
            "https_port": https_result,
            "check_time": datetime.now().isoformat()
        }
    
    @classmethod
    async def _ping_check(cls, ip: str, timeout: int = 3, count: int = 1) -> Dict[str, Any]:
        """
        跨平台Ping检测
        """
        try:
            # 根据操作系统选择ping命令参数
            system = platform.system().lower()
            if system == 'windows':
                cmd = ['ping', '-n', str(count), '-w', str(timeout * 1000), ip]
            else:  # Linux/macOS
                cmd = ['ping', '-c', str(count), '-W', str(timeout), ip]
            
            # 执行ping命令
            result = await asyncio.create_subprocess_exec(
                *cmd,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE
            )
            stdout, stderr = await result.communicate()
            
            success = result.returncode == 0
            output = stdout.decode() if stdout else stderr.decode()
            
            return {
                "success": success,
                "method": "ping",
                "response_time": cls._extract_ping_time(output) if success else None,
                "output": output if not success else None,
                "error": None if success else f"Ping failed (code: {result.returncode})"
            }
        except Exception as e:
            return {
                "success": False,
                "method": "ping",
                "response_time": None,
                "error": str(e)
            }
    
    @classmethod
    async def _tcp_port_check(cls, ip: str, port: int, timeout: int = 3) -> Dict[str, Any]:
        """
        TCP端口连通性检测
        """
        try:
            # 创建连接并设置超时
            future = asyncio.open_connection(ip, port)
            reader, writer = await asyncio.wait_for(future, timeout=timeout)
            
            # 关闭连接
            writer.close()
            await writer.wait_closed()
            
            return {
                "success": True,
                "method": f"tcp_port_{port}",
                "port": port,
                "error": None
            }
        except asyncio.TimeoutError:
            return {
                "success": False,
                "method": f"tcp_port_{port}",
                "port": port,
                "error": f"Connection timeout ({timeout}s)"
            }
        except Exception as e:
            return {
                "success": False,
                "method": f"tcp_port_{port}",
                "port": port,
                "error": str(e)
            }
    
    @classmethod
    def _extract_ping_time(cls, ping_output: str) -> str:
        """
        从ping输出中提取响应时间（跨平台）
        """
        try:
            import re
            # Windows格式：时间=1ms 或 time=1ms  
            # Linux/macOS格式：time=1.23 ms
            patterns = [
                r'时间[=<](\d+\.?\d*)ms',  # Windows中文
                r'time[=<](\d+\.?\d*)\s*ms',  # Linux/macOS英文
                r'time=(\d+\.?\d*)\s*ms',  # 通用格式
            ]
            
            for pattern in patterns:
                match = re.search(pattern, ping_output, re.IGNORECASE)
                if match:
                    return f"{match.group(1)}ms"
        except:
            pass
        return None

    @classmethod
    async def batch_check_business_connectivity_services(
        cls,
        db: AsyncSession,
        device_ids: List[int] = None,
        max_concurrent: int = 20
    ) -> Dict[str, Any]:
        """
        批量检测设备业务IP连通性
        
        Args:
            db: 数据库会话
            device_ids: 设备ID列表，为空则检测所有设备
            max_concurrent: 最大并发数
            
        Returns:
            Dict[str, Any]: 批量检测结果
        """
        from sqlalchemy import select
        
        if device_ids:
            devices_query = select(DeviceInfoDO).where(DeviceInfoDO.device_id.in_(device_ids))
        else:
            devices_query = select(DeviceInfoDO).where(DeviceInfoDO.business_ip.isnot(None))
        
        result = await db.execute(devices_query)
        devices = result.scalars().all()
        
        if not devices:
            return {
                "total_devices": 0,
                "online_devices": 0,
                "offline_devices": 0,
                "check_time": datetime.now().isoformat(),
                "details": []
            }
        
        # 使用信号量限制并发数
        semaphore = asyncio.Semaphore(max_concurrent)
        
        async def check_single_device(device):
            async with semaphore:
                connectivity = await cls.check_business_ip_connectivity_services(
                    db, device_id=device.device_id
                )
                return {
                    "device_id": device.device_id,
                    "hostname": device.hostname,
                    "business_ip": device.business_ip,
                    "online": connectivity["online"],
                    "details": connectivity
                }
        
        # 并发执行所有检测
        logger.info(f"开始批量检测 {len(devices)} 台设备的业务IP连通性")
        
        tasks = [check_single_device(device) for device in devices]
        results = await asyncio.gather(*tasks)
        
        # 统计结果
        online_count = sum(1 for result in results if result["online"])
        offline_count = len(results) - online_count
        
        logger.info(f"批量连通性检测完成: {online_count} 在线, {offline_count} 离线")
        
        return {
            "total_devices": len(devices),
            "online_devices": online_count,
            "offline_devices": offline_count,
            "check_time": datetime.now().isoformat(),
            "details": results
        } 