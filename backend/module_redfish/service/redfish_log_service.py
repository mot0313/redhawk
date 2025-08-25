"""
Redfish日志服务层
"""
import asyncio
from datetime import datetime, timedelta
from typing import List, Optional, Dict, Any
from sqlalchemy.ext.asyncio import AsyncSession
from module_redfish.dao.redfish_log_dao import RedfishLogDao
from module_redfish.dao.device_dao import DeviceDao
from module_redfish.entity.do.redfish_log_do import RedfishLogDO
from module_redfish.entity.vo.redfish_log_vo import (
    RedfishLogQueryModel, AddRedfishLogModel, RedfishLogModel, 
    RedfishLogDetailModel, RedfishLogStatsModel, DeviceLogCollectModel,
    RedfishLogCollectResultModel, RedfishLogCleanupResultModel,
    RedfishLogTempCollectResultModel
)
from module_redfish.core.redfish_client import RedfishClient, decrypt_password
from module_admin.entity.vo.common_vo import CrudResponseModel
from utils.response_util import ResponseUtil
from utils.page_util import PageResponseModel, PageUtil
from utils.log_util import logger


def is_valid_timestamp(timestamp_str: str) -> bool:
    """
    检查时间戳是否有效
    
    Args:
        timestamp_str: 时间戳字符串
        
    Returns:
        bool: 时间戳是否有效
    """
    if not timestamp_str:
        return False
    
    # 检查常见的无效时间格式
    invalid_patterns = [
        "0000-00-00",
        "1900-01-01", 
        "1970-01-01T00:00:00"
    ]
    
    for pattern in invalid_patterns:
        if timestamp_str.startswith(pattern):
            return False
    
    try:
        # 尝试解析时间
        from datetime import datetime
        parsed_time = datetime.fromisoformat(timestamp_str.replace("Z", "+00:00"))
        # 检查年份是否合理（1990年以后）
        if parsed_time.year < 1990:
            return False
        return True
    except:
        return False


class RedfishLogService:
    """Redfish日志服务"""
    
    @classmethod
    async def get_redfish_log_list_services(cls, db: AsyncSession, query_object: RedfishLogQueryModel, is_page: bool = False):
        """
        获取日志列表
        
        Args:
            db: 数据库会话
            query_object: 查询条件
            is_page: 是否分页
            
        Returns:
            分页结果或列表
        """
        try:
            # 获取日志列表
            logs_list, total = await RedfishLogDao.get_redfish_log_list(db, query_object, is_page)
            
            # 转换为响应模型
            logs_vo_list = []
            for log_do in logs_list:
                log_vo = cls._convert_do_to_vo(log_do)
                logs_vo_list.append(log_vo)
            
            if is_page:
                # 手动构建分页响应
                from utils.page_util import PageResponseModel
                return PageResponseModel(
                    rows=logs_vo_list,
                    page_num=query_object.page_num,
                    page_size=query_object.page_size,
                    total=total,
                    has_next=(query_object.page_num * query_object.page_size) < total
                )
            else:
                return logs_vo_list
                
        except Exception as e:
            logger.error(f"获取日志列表失败: {str(e)}")
            raise Exception(f"获取日志列表失败: {str(e)}")
    
    @classmethod
    async def get_redfish_log_detail_services(cls, db: AsyncSession, log_id: str) -> RedfishLogDetailModel:
        """
        获取日志详情
        
        Args:
            db: 数据库会话
            log_id: 日志ID
            
        Returns:
            日志详情
        """
        try:
            log_do = await RedfishLogDao.get_redfish_log_detail(db, log_id)
            if not log_do:
                raise ValueError("日志不存在")
            
            # 获取设备信息以获取主机名
            from module_redfish.dao.device_dao import DeviceDao
            device_do = await DeviceDao.get_device_by_id(db, log_do.device_id)
            hostname = device_do.hostname if device_do else None
            
            return cls._convert_do_to_detail_vo(log_do, hostname)
            
        except ValueError:
            raise
        except Exception as e:
            logger.error(f"获取日志详情失败: {str(e)}")
            raise Exception(f"获取日志详情失败: {str(e)}")
    
    @classmethod
    async def get_redfish_log_stats_services(cls, db: AsyncSession) -> RedfishLogStatsModel:
        """
        获取日志统计信息
        
        Args:
            db: 数据库会话
            
        Returns:
            统计信息
        """
        try:
            stats_data = await RedfishLogDao.get_redfish_log_stats(db)
            return RedfishLogStatsModel(**stats_data)
            
        except Exception as e:
            logger.error(f"获取日志统计信息失败: {str(e)}")
            raise Exception(f"获取日志统计信息失败: {str(e)}")
    
    @classmethod
    async def collect_device_logs_services(cls, db: AsyncSession, 
                                          collect_request: DeviceLogCollectModel,
                                          operator: str):
        """
        收集设备日志
        
        Args:
            db: 数据库会话
            collect_request: 收集请求
            operator: 操作者
            
        Returns:
            收集结果（根据no_storage参数返回不同类型）
        """
        try:
            # 临时收集模式：只支持单设备收集
            if collect_request.no_storage:
                if not collect_request.device_id:
                    return RedfishLogTempCollectResultModel(
                        success=False,
                        message="临时收集模式必须指定设备ID"
                    )
                
                # 获取设备信息
                device = await DeviceDao.get_device_by_id(db, collect_request.device_id)
                if not device:
                    return RedfishLogTempCollectResultModel(
                        success=False,
                        message="未找到指定设备"
                    )
                
                # 临时收集单设备日志
                return await cls._collect_single_device_logs_temp(
                    device=device,
                    log_type=collect_request.log_type,
                    max_entries=collect_request.max_entries,
                    force_refresh=collect_request.force_refresh
                )
            
            # 常规收集模式（原有逻辑）
            # 获取要收集的设备列表
            if collect_request.device_id:
                devices = await DeviceDao.get_device_by_id(db, collect_request.device_id)
                device_list = [devices] if devices else []
            else:
                device_list = await DeviceDao.get_monitoring_devices(db)
            
            if not device_list:
                return RedfishLogCollectResultModel(
                    success=False,
                    device_count=0,
                    total_collected=0,
                    critical_collected=0,
                    warning_collected=0,
                    failed_devices=[],
                    message="没有找到可收集的设备"
                )
            
            # 并发收集日志
            collect_tasks = []
            for device in device_list:
                task = cls._collect_single_device_logs(
                    db=db,
                    device=device,
                    log_type=collect_request.log_type,
                    max_entries=collect_request.max_entries,
                    force_refresh=collect_request.force_refresh,
                    operator=operator
                )
                collect_tasks.append(task)
            
            # 等待所有任务完成
            collect_results = await asyncio.gather(*collect_tasks, return_exceptions=True)
            
            # 汇总结果
            device_count = len(device_list)
            total_collected = 0
            critical_collected = 0
            warning_collected = 0
            failed_devices = []
            
            for i, result in enumerate(collect_results):
                if isinstance(result, Exception):
                    failed_devices.append(device_list[i].oob_ip)
                    logger.error(f"收集设备 {device_list[i].oob_ip} 日志失败: {str(result)}")
                else:
                    total_collected += result.get('total', 0)
                    critical_collected += result.get('critical', 0)
                    warning_collected += result.get('warning', 0)
            
            success = len(failed_devices) < device_count
            message = f"成功收集 {device_count - len(failed_devices)} 台设备的日志，失败 {len(failed_devices)} 台"
            
            return RedfishLogCollectResultModel(
                success=success,
                device_count=device_count,
                total_collected=total_collected,
                critical_collected=critical_collected,
                warning_collected=warning_collected,
                failed_devices=failed_devices,
                message=message
            )
            
        except Exception as e:
            logger.error(f"收集设备日志失败: {str(e)}")
            raise Exception(f"收集设备日志失败: {str(e)}")
    
    @classmethod
    async def _collect_single_device_logs(cls, db: AsyncSession, device, log_type: str, 
                                         max_entries: int, force_refresh: bool, operator: str) -> Dict[str, int]:
        """
        收集单个设备的日志
        
        Args:
            db: 数据库会话
            device: 设备对象
            log_type: 日志类型
            max_entries: 最大条目数
            force_refresh: 是否强制刷新
            operator: 操作者
            
        Returns:
            收集结果字典
        """
        try:
            # 解密密码
            try:
                decrypted_password = decrypt_password(device.redfish_password)
            except ValueError as e:
                # 密码解密失败，抛出友好的错误信息
                raise Exception(f"设备 {device.oob_ip} 密码解密失败: {str(e)}")
            
            # 创建Redfish客户端
            client = RedfishClient(
                host=device.oob_ip,
                username=device.redfish_username,
                password=decrypted_password,
                port=device.oob_port or 443,
                timeout=30
            )
            
            # 获取最新日志时间（如果不是强制刷新）
            since_timestamp = None
            if not force_refresh:
                since_timestamp = await RedfishLogDao.get_latest_log_time(db, device.device_id)
            
            # 获取日志
            logs_data = await client.get_event_logs(
                log_type=log_type,
                max_entries=max_entries,
                since_timestamp=since_timestamp
            )
            
            # 过滤并保存日志
            saved_count = 0
            critical_count = 0
            warning_count = 0
            logs_to_save = []
            
            for log_entry in logs_data:
                # 检查是否已存在
                entry_id = log_entry.get('id', '')
                created_time_str = log_entry.get('created', '')
                
                if not entry_id:
                    continue
                
                # 处理创建时间
                created_time = None
                is_time_valid = False
                
                if created_time_str and is_valid_timestamp(created_time_str):
                    try:
                        # 解析有效时间并转换为无时区的datetime
                        parsed_time = datetime.fromisoformat(created_time_str.replace("Z", "+00:00"))
                        created_time = parsed_time.replace(tzinfo=None)
                        is_time_valid = True
                    except:
                        pass
                
                # 对于无效时间，使用一个固定的特殊时间戳表示"未知时间"
                if created_time is None:
                    # 使用1900-01-01作为"未知时间"的标识
                    created_time = datetime(1900, 1, 1)
                    is_time_valid = False
                
                # 检查是否已存在
                exists = await RedfishLogDao.check_log_exists(db, device.device_id, entry_id, created_time)
                if exists:
                    continue
                
                # 创建日志对象
                # 将完整的原始Redfish响应保存到备注中
                import json
                
                # 只保存完整的Redfish响应
                original_log_info = log_entry.get('raw_data', {})
                
                log_model = AddRedfishLogModel(
                    device_id=device.device_id,
                    device_ip=device.oob_ip,
                    entry_id=entry_id,
                    entry_type=log_entry.get('entry_type', ''),
                    log_source=log_entry.get('log_source', ''),
                    severity=log_entry.get('severity', ''),
                    created_time=created_time,
                    message=log_entry.get('message', ''),
                    remark=json.dumps(original_log_info, ensure_ascii=False, indent=2),
                    create_by=operator,
                    create_time=datetime.now()
                )
                
                logs_to_save.append(log_model)
                saved_count += 1
                
                if log_entry.get('severity', '').upper() == 'CRITICAL':
                    critical_count += 1
                elif log_entry.get('severity', '').upper() == 'WARNING':
                    warning_count += 1
            
            # 批量保存日志
            if logs_to_save:
                await RedfishLogDao.add_redfish_logs_batch(db, logs_to_save)
                await db.commit()
            
            return {
                'total': saved_count,
                'critical': critical_count,
                'warning': warning_count
            }
            
        except Exception as e:
            import traceback
            error_msg = str(e) if str(e) else f"未知错误: {type(e).__name__}"
            error_traceback = traceback.format_exc()
            logger.error(f"收集设备 {device.oob_ip} 日志失败: {error_msg}")
            logger.error(f"错误详情: {error_traceback}")
            raise Exception(f"收集设备 {device.oob_ip} 日志失败: {error_msg}")
    
    @classmethod
    async def _collect_single_device_logs_temp(cls, device, log_type: str, 
                                             max_entries: int, force_refresh: bool) -> RedfishLogTempCollectResultModel:
        """
        临时收集单个设备的日志（不保存到数据库）
        
        Args:
            device: 设备对象
            log_type: 日志类型
            max_entries: 最大条目数
            force_refresh: 是否强制刷新
            
        Returns:
            临时收集结果
        """
        try:
            # 解密密码
            try:
                decrypted_password = decrypt_password(device.redfish_password)
            except ValueError as e:
                # 密码解密失败，返回友好的错误信息
                return RedfishLogTempCollectResultModel(
                    success=False,
                    device_id=device.device_id,
                    device_ip=device.oob_ip,
                    device_name=device.hostname,
                    message=f"密码解密失败: {str(e)}"
                )
            
            # 创建Redfish客户端
            client = RedfishClient(
                host=device.oob_ip,
                username=device.redfish_username,
                password=decrypted_password,
                port=device.oob_port or 443,
                timeout=30
            )
            
            # 获取日志（临时模式下不使用since_timestamp）
            logs_data = await client.get_event_logs(
                log_type=log_type,
                max_entries=max_entries,
                since_timestamp=None
            )
            
            # 过滤和统计日志
            critical_count = 0
            warning_count = 0
            filtered_logs = []
            
            for log_entry in logs_data:
                severity = log_entry.get('severity', '').upper()
                
                # 只处理Critical和Warning级别的日志
                if severity in ['CRITICAL', 'WARNING']:
                    # 处理创建时间
                    created_time_str = log_entry.get('created', '')
                    is_time_valid = is_valid_timestamp(created_time_str) if created_time_str else False
                    
                    # 生成完整的原始日志信息JSON
                    import json
                    
                    # 只保存完整的Redfish响应
                    original_log_info = log_entry.get('raw_data', {})
                    
                    # 转换日志格式供前端显示
                    log_display = {
                        'entryId': log_entry.get('id', ''),
                        'entryType': log_entry.get('entry_type', ''),
                        'logSource': log_entry.get('log_source', ''),
                        'severity': severity,
                        'createdTime': log_entry.get('created', ''),
                        'message': log_entry.get('message', ''),
                        'originalData': json.dumps(original_log_info, ensure_ascii=False, indent=2)
                    }
                    
                    filtered_logs.append(log_display)
                    
                    if severity == 'CRITICAL':
                        critical_count += 1
                    elif severity == 'WARNING':
                        warning_count += 1
            
            return RedfishLogTempCollectResultModel(
                success=True,
                device_id=device.device_id,
                device_ip=device.oob_ip,
                device_name=device.hostname,
                total_collected=len(filtered_logs),
                critical_count=critical_count,
                warning_count=warning_count,
                logs_data=filtered_logs,
                message=f"成功收集设备 {device.oob_ip} 的 {len(filtered_logs)} 条日志"
            )
            
        except Exception as e:
            import traceback
            error_msg = str(e) if str(e) else f"未知错误: {type(e).__name__}"
            error_traceback = traceback.format_exc()
            logger.error(f"临时收集设备 {device.oob_ip} 日志失败: {error_msg}")
            logger.error(f"错误详情: {error_traceback}")
            return RedfishLogTempCollectResultModel(
                success=False,
                device_id=device.device_id,
                device_ip=device.oob_ip,
                device_name=device.hostname,
                message=f"收集失败: {error_msg}"
            )
    
    @classmethod
    async def cleanup_old_logs_services(cls, db: AsyncSession, days: int = 30) -> RedfishLogCleanupResultModel:
        """
        清理旧日志
        
        Args:
            db: 数据库会话
            days: 保留天数
            
        Returns:
            清理结果
        """
        try:
            before_date = datetime.now() - timedelta(days=days)
            cleaned_count = await RedfishLogDao.cleanup_old_logs(db, before_date)
            await db.commit()
            
            return RedfishLogCleanupResultModel(
                success=True,
                cleaned_count=cleaned_count,
                before_date=before_date,
                message=f"成功清理 {cleaned_count} 条 {days} 天前的日志"
            )
            
        except Exception as e:
            logger.error(f"清理旧日志失败: {str(e)}")
            await db.rollback()
            raise Exception(f"清理旧日志失败: {str(e)}")
    
    @classmethod
    async def delete_redfish_log_services(cls, db: AsyncSession, log_id: str) -> CrudResponseModel:
        """
        删除日志
        
        Args:
            db: 数据库会话
            log_id: 日志ID
            
        Returns:
            操作结果
        """
        try:
            # 检查日志是否存在
            log_obj = await RedfishLogDao.get_redfish_log_detail(db, log_id)
            if not log_obj:
                return ResponseUtil.failure(msg="日志不存在")
            
            # 删除日志
            success = await RedfishLogDao.delete_redfish_log(db, log_id)
            if success:
                await db.commit()
                return ResponseUtil.success(msg="删除成功")
            else:
                return ResponseUtil.failure(msg="删除失败")
                
        except Exception as e:
            logger.error(f"删除日志失败: {str(e)}")
            await db.rollback()
            return ResponseUtil.failure(msg="删除失败")
    
    @classmethod
    async def delete_device_logs_services(cls, db: AsyncSession, device_id: int) -> CrudResponseModel:
        """
        删除指定设备的所有日志
        
        Args:
            db: 数据库会话
            device_id: 设备ID
            
        Returns:
            操作结果
        """
        try:
            deleted_count = await RedfishLogDao.delete_redfish_logs_by_device(db, device_id)
            await db.commit()
            return ResponseUtil.success(data={"deleted_count": deleted_count}, msg=f"成功删除 {deleted_count} 条日志")
            
        except Exception as e:
            logger.error(f"删除设备日志失败: {str(e)}")
            await db.rollback()
            return ResponseUtil.failure(msg="删除设备日志失败")
    
    @classmethod
    def _convert_do_to_vo(cls, log_do: RedfishLogDO, hostname: str = None) -> RedfishLogModel:
        """将数据对象转换为视图对象"""
        return RedfishLogModel(
            log_id=str(log_do.log_id),
            device_id=log_do.device_id,
            device_ip=log_do.device_ip,
            hostname=hostname,
            entry_id=log_do.entry_id,
            entry_type=log_do.entry_type,
            log_source=log_do.log_source,
            severity=log_do.severity,
            created_time=log_do.created_time,
            collected_time=log_do.collected_time,
            message=log_do.message,
            remark=log_do.remark
        )
    
    @classmethod
    def _convert_do_to_detail_vo(cls, log_do: RedfishLogDO, hostname: str = None) -> RedfishLogDetailModel:
        """将数据对象转换为详细视图对象"""
        return RedfishLogDetailModel(
            log_id=str(log_do.log_id),
            device_id=log_do.device_id,
            device_ip=log_do.device_ip,
            hostname=hostname,
            entry_id=log_do.entry_id,
            entry_type=log_do.entry_type,
            log_source=log_do.log_source,
            severity=log_do.severity,
            created_time=log_do.created_time,
            collected_time=log_do.collected_time,
            message=log_do.message,
            remark=log_do.remark,
            create_by=log_do.create_by,
            create_time=log_do.create_time,
            update_by=log_do.update_by,
            update_time=log_do.update_time
        )
    
    @staticmethod
    async def export_logs_data_services(logs_list: List):
        """
        导出日志数据service
        
        Args:
            logs_list: 日志列表
            
        Returns:
            日志数据对应excel的二进制数据
        """
        import io
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        
        # 创建映射字典，将英文键映射到中文键
        mapping_dict = {
            'serialNo': '序号',
            'deviceIp': '设备IP',
            'entryId': '条目ID',
            'entryType': '条目类型',
            'logSource': '日志来源',
            'severity': '严重程度',
            'createdTime': '创建时间',
            'collectedTime': '收集时间',
            'message': '消息内容',
            'remark': '备注'
        }
        
        # 处理数据格式化
        processed_data = []
        for index, log in enumerate(logs_list, 1):
            # 确保log是字典类型
            if not isinstance(log, dict):
                if hasattr(log, 'model_dump'):
                    log_dict = log.model_dump(by_alias=True)
                elif hasattr(log, '__dict__'):
                    log_dict = {}
                    for key, value in log.__dict__.items():
                        if not key.startswith('_'):
                            from utils.common_util import CamelCaseUtil
                            camel_key = CamelCaseUtil.snake_to_camel(key)
                            log_dict[camel_key] = value
                    log_dict = log_dict
                else:
                    continue
            else:
                log_dict = log
            
            # 格式化严重程度
            severity_map = {
                'CRITICAL': '严重',
                'WARNING': '警告',
                'OK': '正常',
                'INFO': '信息'
            }
            severity = log_dict.get('severity')
            if severity:
                log_dict['severity'] = severity_map.get(severity, severity)
            
            # 格式化时间
            time_fields = ['createdTime', 'collectedTime']
            for time_field in time_fields:
                time_value = log_dict.get(time_field)
                if time_value:
                    try:
                        if isinstance(time_value, str):
                            log_dict[time_field] = time_value
                        else:
                            log_dict[time_field] = time_value.strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        log_dict[time_field] = ''
            
            # 添加序号
            log_dict['serialNo'] = index
            processed_data.append(log_dict)
        
        # 创建Excel文件
        wb = Workbook()
        ws = wb.active
        ws.title = "日志列表"
        
        # 表头样式设置
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        # 获取表头列表
        headers = list(mapping_dict.values())
        keys = list(mapping_dict.keys())
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(col_num)].width = 15
        
        # 写入数据
        for row_num, item in enumerate(processed_data, 2):
            for col_num, key in enumerate(keys, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = item.get(key, '')
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 保存为字节数据
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        return file_stream.getvalue()
